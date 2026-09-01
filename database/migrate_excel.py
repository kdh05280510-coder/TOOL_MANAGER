"""
이관만 할 때

1. exe를 끈다
2. VS Code 또는 검은 창에서 이관 스크립트 실행
    예: python import_excel.py
3. 같은 폴더의 DB가 갱신된다
4. 이미 만들어 둔 공구등록_VER2.2.1.exe를 다시 연다

exe와 DB를 같은 폴더에 두면 됩니다. 예: D:\tool_manager\dist\ 에 exe와 tool_manager.db가 함께 있는 형태입니다.
코드가 안 바뀌었으면 python main.py로 실행해도 되고, 예전 exe를 그대로 써도 됩니다.
"""

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from database.db import get_connection, init_db

# 엑셀 파일 후보
CANDIDATES = [
    ROOT / "공구관리대장_성진세미텍.xlsm",
    ROOT / "공구관리대장_성진세미텍ver2.xlsm",
]

SHEETS = {
    "EM(ALU)": 0,
    "EM(SUS)": 0,
    "EM(STEEL)": 0,
    "DR": 0,
    "SP": 0,
    "B급 재등록": 1,
}


def s(v):
    if v is None:
        return ""
    return str(v).strip()


def to_float(v):
    text = s(v)
    if not text or text == "None":
        return None
    text = text.replace("생크직경:", "").replace("전체길이:", "").replace("온길이:", "")
    try:
        return float(text)
    except ValueError:
        m = re.search(r"([0-9.]+)", text)
        return float(m.group(1)) if m else None


def parse_dl(tool_name):
    text = s(tool_name)
    d = l = None
    m = re.search(r"[Dd]\s*([0-9.]+)", text)
    if m:
        d = float(m.group(1))
    m = re.search(r"[Ll]\s*([0-9.]+)", text)
    if m:
        l = float(m.group(1))
    return d, l


def get_or_create_maker(cur, name):
    name = s(name)
    if not name or name in ("-", "?", "None"):
        return None
    cur.execute("SELECT id FROM makers WHERE name = ?", (name,))
    row = cur.fetchone()
    if row:
        return row["id"]
    cur.execute("INSERT INTO makers (name, is_active) VALUES (?, 1)", (name,))
    return cur.lastrowid


def get_or_create_category(cur, main_name, sub_code):
    main_name = s(main_name)
    sub_code = s(sub_code)
    if not main_name or not sub_code or main_name == "None" or sub_code == "None":
        return None

    cur.execute(
        "SELECT id FROM categories WHERE main_name=? AND sub_code=?",
        (main_name, sub_code),
    )
    row = cur.fetchone()
    if row:
        return row["id"]

    # N(엔드밀) → N , DR(드릴) → DR
    main_code = main_name
    if "(" in main_name:
        main_code = main_name.split("(")[0].strip()
    if not main_code:
        main_code = main_name

    try:
        cur.execute(
            "INSERT INTO categories (main_code, main_name, sub_code) VALUES (?, ?, ?)",
            (main_code, main_name, sub_code),
        )
    except Exception:
        # 컬럼 구성이 조금 다를 때 대비
        cur.execute(
            "INSERT INTO categories (main_code, main_name, sub_code, is_active) VALUES (?, ?, ?, 1)",
            (main_code, main_name, sub_code),
        )

    print("  분류 추가:", main_code, main_name, "/", sub_code)
    return cur.lastrowid


def find_excel():
    for p in CANDIDATES:
        if p.exists():
            return p
    return None


def migrate():
    init_db()
    path = find_excel()
    if not path:
        print("엑셀 파일을 찾을 수 없습니다. 아래 중 하나를 두세요:")
        for p in CANDIDATES:
            print(" -", p)
        return

    print("엑셀:", path)
    wb = load_workbook(path, data_only=True)
    conn = get_connection()
    cur = conn.cursor()

    tools_ok = 0
    inv_ok = 0
    skip = 0

    for sheet_name, is_b in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print("시트 없음:", sheet_name)
            continue

        ws = wb[sheet_name]
        print("처리 중:", sheet_name)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            # A번호 B대분류 C소분류 D제조사 E상품명 F상품명(부) G상품코드 H바코드 I수량 J생크 K전체길이
            main_name = row[1] if len(row) > 1 else None
            sub_code = row[2] if len(row) > 2 else None
            maker_name = row[3] if len(row) > 3 else None
            tool_name = row[4] if len(row) > 4 else None
            sub_name = row[5] if len(row) > 5 else None
            tool_code = row[6] if len(row) > 6 else None
            barcode = row[7] if len(row) > 7 else None
            shank = row[9] if len(row) > 9 else None
            total = row[10] if len(row) > 10 else None

            barcode = s(barcode)
            if not barcode or barcode == "None":
                skip += 1
                continue

            cur.execute("SELECT id FROM inventory WHERE barcode = ?", (barcode,))
            if cur.fetchone():
                skip += 1
                continue

            cat_id = get_or_create_category(cur, main_name, sub_code)
            if not cat_id:
                print("  분류 비어있음 스킵:", main_name, sub_code, barcode)
                skip += 1
                continue

            maker_id = get_or_create_maker(cur, maker_name)
            diameter, length = parse_dl(tool_name)

            cur.execute("""
                INSERT INTO tools (
                    category_id, maker_id, tool_code, tool_name,
                    diameter, length, shank_dia, total_length
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                cat_id,
                maker_id,
                s(tool_code) or None,
                s(tool_name) or None,
                diameter,
                length,
                to_float(shank),
                to_float(total),
            ))
            tool_id = cur.lastrowid
            tools_ok += 1

            cur.execute("""
                INSERT INTO inventory (
                    tool_id, barcode, sub_name, quantity, status, is_grade_b
                ) VALUES (?, ?, ?, 1, ?, ?)
            """, (
                tool_id,
                barcode,
                s(sub_name) or None,
                "B급" if is_b else "정상",
                1 if is_b else 0,
            ))
            inv_ok += 1

        conn.commit()

    conn.close()
    print("이관 완료")
    print("tools 추가:", tools_ok)
    print("inventory 추가:", inv_ok)
    print("스킵:", skip)


if __name__ == "__main__":
    migrate()